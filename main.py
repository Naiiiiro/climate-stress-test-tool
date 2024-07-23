#%%
import os
import pandas as pd
from openpyxl.styles import Alignment
from modules.credit_risk_assessment.calculate_pd import process_PD_domestic_corporate_credit, process_PD_domestic_personal_mortgage, process_PD_domestic_personal_other, process_PD_overseas_credit, process_PD_domestic_investment, process_PD_overseas_investment
from modules.credit_risk_assessment.calculate_lgd import process_LGD_domestic_corporate_credit, process_LGD_domestic_personal_mortgage, process_LGD_domestic_personal_other, process_LGD_overseas_credit, process_LGD_domestic_investment, process_LGD_overseas_investment
from modules.credit_risk_assessment.calculate_ead import process_EAD_domestic_corporate_credit, process_EAD_domestic_personal_mortgage, process_EAD_domestic_personal_other, process_EAD_overseas_credit, process_EAD_domestic_investment, process_EAD_overseas_investment

def calculate_expected_loss(ead, pd_rate, lgd_rate):
    """Calculate the expected loss given EAD, PD rate, and LGD rate.

    Args:
        ead (float): Exposure at default.
        pd_rate (float or list): Probability of default rate.
        lgd_rate (float or list): Loss given default rate.

    Returns:
        float: Expected loss.
    """
    if isinstance(pd_rate, list):
        pd_rate = pd_rate[0]
    if isinstance(lgd_rate, list):
        lgd_rate = lgd_rate[0]
    el = pd_rate / 100 * lgd_rate / 100 * ead * 100
    return el


#%%
def save_to_excel(summary_df, sheet_name, writer, pd_rate, lgd_rate, pd_rate_orderly, lgd_rate_orderly, pd_rate_disorderly, lgd_rate_disorderly, pd_rate_no_policy, lgd_rate_no_policy, ead):
    """Save the summary data to an Excel sheet with proper formatting.

    Args:
        summary_df (pd.DataFrame): Summary data frame.
        sheet_name (str): Name of the sheet to save data.
        writer (pd.ExcelWriter): Excel writer object.
        pd_rate (list): List of PD rates.
        lgd_rate (list): List of LGD rates.
        pd_rate_orderly (list): List of orderly transition PD rates.
        lgd_rate_orderly (list): List of orderly transition LGD rates.
        pd_rate_disorderly (list): List of disorderly transition PD rates.
        lgd_rate_disorderly (list): List of disorderly transition LGD rates.
        pd_rate_no_policy (list): List of no policy scenario PD rates.
        lgd_rate_no_policy (list): List of no policy scenario LGD rates.
        ead (list): List of EAD values.
    """
    ead_df = pd.DataFrame(ead)
    summary_df = summary_df.dropna(subset=['客戶名'])
    summary_df = summary_df.merge(ead_df, on='客戶名', how='left')
    summary_df.rename(columns={'EAD': '曝險金額'}, inplace=True)

    pd_rate_df = pd.DataFrame(pd_rate)
    lgd_rate_df = pd.DataFrame(lgd_rate)
    pd_rate_orderly_df = pd.DataFrame(pd_rate_orderly)
    lgd_rate_orderly_df = pd.DataFrame(lgd_rate_orderly)
    pd_rate_disorderly_df = pd.DataFrame(pd_rate_disorderly)
    lgd_rate_disorderly_df = pd.DataFrame(lgd_rate_disorderly)
    pd_rate_no_policy_df = pd.DataFrame(pd_rate_no_policy)
    lgd_rate_no_policy_df = pd.DataFrame(lgd_rate_no_policy)

    summary_data = []

    for idx, row in summary_df.iterrows():
        company = row['客戶名']
        company_ead = row['曝險金額']

        # Extract PD and LGD rates for different scenarios
        company_pd_rate = pd_rate_df[pd_rate_df['客戶名'] == company]['PD(%)'].values
        company_lgd_rate = lgd_rate_df[lgd_rate_df['客戶名'] == company]['LGD(%)'].values
        company_pd_rate_orderly = pd_rate_orderly_df[pd_rate_orderly_df['客戶名'] == company]['PD(%)'].values
        company_lgd_rate_orderly = lgd_rate_orderly_df[lgd_rate_orderly_df['客戶名'] == company]['LGD(%)'].values
        company_pd_rate_disorderly = pd_rate_disorderly_df[pd_rate_disorderly_df['客戶名'] == company]['PD(%)'].values
        company_lgd_rate_disorderly = lgd_rate_disorderly_df[lgd_rate_disorderly_df['客戶名'] == company]['LGD(%)'].values
        company_pd_rate_no_policy = pd_rate_no_policy_df[pd_rate_no_policy_df['客戶名'] == company]['PD(%)'].values
        company_lgd_rate_no_policy = lgd_rate_no_policy_df[lgd_rate_no_policy_df['客戶名'] == company]['LGD(%)'].values

        result = {
            '客戶名': company,
            '曝險金額': company_ead,
            '基準情境_平均違約率': float('nan'), '基準情境_平均違約損失率': float('nan'), '基準情境_估計可能損失數': float('nan'), 
            '有序轉型_平均違約率': float('nan'), '有序轉型_平均違約損失率': float('nan'), '有序轉型_估計可能損失數': float('nan'), 
            '無序轉型_平均違約率': float('nan'), '無序轉型_平均違約損失率': float('nan'), '無序轉型_估計可能損失數': float('nan'), 
            '無政策情境_平均違約率': float('nan'), '無政策情境_平均違約損失率': float('nan'), '無政策情境_估計可能損失數': float('nan')
        }

        # Baseline scenario
        if len(company_pd_rate) > 0:
            result['基準情境_平均違約率'] = company_pd_rate[0]
            result['基準情境_估計可能損失數'] = float('nan')
        if len(company_lgd_rate) > 0:
            result['基準情境_平均違約損失率'] = company_lgd_rate[0] *100
            result['基準情境_估計可能損失數'] = float('nan')
        if len(company_pd_rate) > 0 and len(company_lgd_rate) > 0:
            base_el = calculate_expected_loss(company_ead, company_pd_rate[0], company_lgd_rate[0])
            result['基準情境_估計可能損失數'] = base_el

        # Orderly transition scenario
        if len(company_pd_rate_orderly) > 0:
            result['有序轉型_平均違約率'] = company_pd_rate_orderly[0]
            result['有序轉型_估計可能損失數'] = float('nan')
        if len(company_lgd_rate_orderly) > 0:
            result['有序轉型_平均違約損失率'] = company_lgd_rate_orderly[0]*100
            result['有序轉型_估計可能損失數'] = float('nan')
        elif len(company_pd_rate_orderly) > 0 and len(company_lgd_rate_orderly) > 0:
            orderly_el = calculate_expected_loss(company_ead, company_pd_rate_orderly[0], company_lgd_rate_orderly[0])
            result['有序轉型_估計可能損失數'] = orderly_el

        # Disorderly transition scenario
        if len(company_pd_rate_disorderly) > 0:
            result['無序轉型_平均違約率'] = company_pd_rate_disorderly[0] 
            result['無序轉型_估計可能損失數'] = float('nan')
        if len(company_lgd_rate_disorderly) > 0:
            result['無序轉型_平均違約損失率'] = company_lgd_rate_disorderly[0]*100
            result['無序轉型_估計可能損失數'] = float('nan')
        if len(company_pd_rate_disorderly) > 0 and len(company_lgd_rate_disorderly) > 0:
            disorderly_el = calculate_expected_loss(company_ead, company_pd_rate_disorderly[0], company_lgd_rate_disorderly[0])
            result['無序轉型_估計可能損失數'] = disorderly_el        

        # No policy scenario
        if len(company_pd_rate_no_policy) > 0:
            result['無政策情境_平均違約率'] = company_pd_rate_no_policy[0]
            result['無政策情境_估計可能損失數'] = float('nan')
        if len(company_lgd_rate_no_policy) > 0:
            result['無政策情境_平均違約損失率'] = company_lgd_rate_no_policy[0]*100
            result['無政策情境_估計可能損失數'] = float('nan')
        if len(company_pd_rate_no_policy) > 0 and len(company_lgd_rate_no_policy) > 0:
            no_policy_el = calculate_expected_loss(company_ead, company_pd_rate_no_policy[0], company_lgd_rate_no_policy[0])
            result['無政策情境_估計可能損失數'] = no_policy_el

        summary_data.append(result)

    summary_result_df = pd.DataFrame(summary_data)
    
    # Update column titles
    summary_result_df.columns = ['客戶名', '曝險金額'] + ['平均違約率(%)', '平均違約損失率(%)', '估計可能損失數'] * 4

    # Save the summary data frame to the specified Excel sheet
    summary_result_df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=1)
    worksheet = writer.sheets[sheet_name]

    # Merge cells for the scenario headers and set alignment
    worksheet.merge_cells('C1:E1')
    worksheet['C1'] = '基準情境'
    worksheet['C1'].alignment = Alignment(horizontal='center', vertical='center')

    worksheet.merge_cells('F1:H1')
    worksheet['F1'] = '有序轉型'
    worksheet['F1'].alignment = Alignment(horizontal='center', vertical='center')

    worksheet.merge_cells('I1:K1')
    worksheet['I1'] = '無序轉型'
    worksheet['I1'].alignment = Alignment(horizontal='center', vertical='center')

    worksheet.merge_cells('L1:N1')
    worksheet['L1'] = '無政策情境'
    worksheet['L1'].alignment = Alignment(horizontal='center', vertical='center')

def save_to_excel_overseas(summary_df, sheet_name, writer, pd_rate, lgd_rate, pd_rate_orderly, lgd_rate_orderly, pd_rate_disorderly, lgd_rate_disorderly, ead):
    """Save the overseas summary data to an Excel sheet with proper formatting.

    Args:
        summary_df (pd.DataFrame): Summary data frame.
        sheet_name (str): Name of the sheet to save data.
        writer (pd.ExcelWriter): Excel writer object.
        pd_rate (list): List of PD rates.
        lgd_rate (list): List of LGD rates.
        pd_rate_orderly (list): List of orderly transition PD rates.
        lgd_rate_orderly (list): List of orderly transition LGD rates.
        pd_rate_disorderly (list): List of disorderly transition PD rates.
        lgd_rate_disorderly (list): List of disorderly transition LGD rates.
        ead (list): List of EAD values.
    """
    ead_df = pd.DataFrame(ead)
    summary_df = summary_df.dropna(subset=['客戶名'])
    summary_df = summary_df.merge(ead_df, on='客戶名', how='left')
    summary_df.rename(columns={'EAD': '曝險金額'}, inplace=True)

    pd_rate_df = pd.DataFrame(pd_rate)
    lgd_rate_df = pd.DataFrame(lgd_rate)
    pd_rate_orderly_df = pd.DataFrame(pd_rate_orderly)
    lgd_rate_orderly_df = pd.DataFrame(lgd_rate_orderly)
    pd_rate_disorderly_df = pd.DataFrame(pd_rate_disorderly)
    lgd_rate_disorderly_df = pd.DataFrame(lgd_rate_disorderly)

    summary_data = []

    for idx, row in summary_df.iterrows():
        company = row['客戶名']
        company_ead = row['曝險金額']

        # Ensure the extracted data is not a list
        company_pd_rate = pd_rate_df[pd_rate_df['客戶名'] == company]['PD(%)'].values
        company_lgd_rate = lgd_rate_df[lgd_rate_df['客戶名'] == company]['LGD(%)'].values
        company_pd_rate_orderly = pd_rate_orderly_df[pd_rate_orderly_df['客戶名'] == company]['PD(%)'].values
        company_lgd_rate_orderly = lgd_rate_orderly_df[lgd_rate_orderly_df['客戶名'] == company]['LGD(%)'].values
        company_pd_rate_disorderly = pd_rate_disorderly_df[pd_rate_disorderly_df['客戶名'] == company]['PD(%)'].values
        company_lgd_rate_disorderly = lgd_rate_disorderly_df[lgd_rate_disorderly_df['客戶名'] == company]['LGD(%)'].values

        result = {
            '客戶名': company,
            '曝險金額': company_ead,
            '基準情境_平均違約率': company_pd_rate[0] if len(company_pd_rate) > 0 else float('nan'),
            '基準情境_平均違約損失率': company_lgd_rate[0]*100 if len(company_lgd_rate) > 0 else float('nan'),
            '基準情境_估計可能損失數': float('nan'),
            '有序轉型_平均違約率': company_pd_rate_orderly[0] if len(company_pd_rate_orderly) > 0 else float('nan'),
            '有序轉型_平均違約損失率': company_lgd_rate_orderly[0]*100 if len(company_lgd_rate_orderly) > 0 else float('nan'),
            '有序轉型_估計可能損失數': float('nan'),
            '無序轉型_平均違約率': company_pd_rate_disorderly[0] if len(company_pd_rate_disorderly) > 0 else float('nan'),
            '無序轉型_平均違約損失率': company_lgd_rate_disorderly[0]*100 if len(company_lgd_rate_disorderly) > 0 else float('nan'),
            '無序轉型_估計可能損失數': float('nan')
        }

        # Baseline scenario
        if not pd.isna(result['基準情境_平均違約率']) and not pd.isna(result['基準情境_平均違約損失率']):
            base_el = calculate_expected_loss(company_ead, result['基準情境_平均違約率'], result['基準情境_平均違約損失率'])
            result['基準情境_估計可能損失數'] = base_el

        # Orderly transition scenario
        if not pd.isna(result['有序轉型_平均違約率']) and not pd.isna(result['有序轉型_平均違約損失率']):
            orderly_el = calculate_expected_loss(company_ead, result['有序轉型_平均違約率'], result['有序轉型_平均違約損失率'])
            result['有序轉型_估計可能損失數'] = orderly_el

        # Disorderly transition scenario
        if not pd.isna(result['無序轉型_平均違約率']) and not pd.isna(result['無序轉型_平均違約損失率']):
            disorderly_el = calculate_expected_loss(company_ead, result['無序轉型_平均違約率'], result['無序轉型_平均違約損失率'])
            result['無序轉型_估計可能損失數'] = disorderly_el        

        summary_data.append(result)

    summary_result_df = pd.DataFrame(summary_data)

    # Update column titles
    summary_result_df.columns = ['客戶名', '曝險金額'] + ['平均違約率(%)', '平均違約損失率(%)', '估計可能損失數'] * 3

    # Save the summary data frame to the specified Excel sheet
    summary_result_df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=1)
    worksheet = writer.sheets[sheet_name]

    # Merge cells for the scenario headers and set alignment
    worksheet.merge_cells('C1:E1')
    worksheet['C1'] = '基準情境'
    worksheet['C1'].alignment = Alignment(horizontal='center', vertical='center')

    worksheet.merge_cells('F1:H1')
    worksheet['F1'] = '有序轉型'
    worksheet['F1'].alignment = Alignment(horizontal='center', vertical='center')

    worksheet.merge_cells('I1:K1')
    worksheet['I1'] = '無序轉型'
    worksheet['I1'].alignment = Alignment(horizontal='center', vertical='center')


def process_domestic_corporate_credit():
    input_dir = 'input-data_files/授信部位/國內授信/'
    output_dir = 'results-data_files/授信部位/國內授信/'

    if not os.path.exists(output_dir):
        os.makedirs(output_dir)

    for filename in os.listdir(input_dir):
        if filename.endswith('國內企業授信.xlsx'):
            file_path = os.path.join(input_dir, filename)
            xls = pd.ExcelFile(file_path)

            pd_rate, pd_rate_orderly_2030, pd_rate_disorderly_2030, pd_rate_no_policy_2030, pd_rate_orderly_2050, pd_rate_disorderly_2050, pd_rate_no_policy_2050, pd_rate_no_policy_2090 = process_PD_domestic_corporate_credit(file_path)
            lgd_rate, lgd_rate_orderly_2030, lgd_rate_disorderly_2030, lgd_rate_no_policy_2030, lgd_rate_orderly_2050, lgd_rate_disorderly_2050, lgd_rate_no_policy_2050, lgd_rate_no_policy_2090 = process_LGD_domestic_corporate_credit(file_path)
            ead = process_EAD_domestic_corporate_credit(file_path)

            summary_2030 = []
            summary_2050 = []
            
            for sheet_name in xls.sheet_names:
                df = pd.read_excel(xls, sheet_name=sheet_name)
                summary_2030.append(df)
                summary_2050.append(df)
            
            summary_df_2030 = pd.concat(summary_2030, ignore_index=True)
            summary_df_2050 = pd.concat(summary_2050, ignore_index=True)

            output_file_path = os.path.join(output_dir, filename)
            with pd.ExcelWriter(output_file_path, engine='openpyxl') as writer:
                save_to_excel(summary_df_2030, '國內授信彙總表(2030年)(企業授信)', writer, pd_rate, lgd_rate, pd_rate_orderly_2030, lgd_rate_orderly_2030, pd_rate_disorderly_2030, lgd_rate_disorderly_2030, pd_rate_no_policy_2030, lgd_rate_no_policy_2030, ead)
                save_to_excel(summary_df_2050, '國內授信彙總表(2050年)(企業授信)', writer, pd_rate, lgd_rate, pd_rate_orderly_2050, lgd_rate_orderly_2050, pd_rate_disorderly_2050, lgd_rate_disorderly_2050, pd_rate_no_policy_2050, lgd_rate_no_policy_2050, ead)

                print(f"結果已儲存至 {output_file_path}")

def process_domestic_personal_mortgage_credit():
    input_dir = 'input-data_files/授信部位/國內授信/'
    output_dir = 'results-data_files/授信部位/國內授信/'

    if not os.path.exists(output_dir):
        os.makedirs(output_dir)

    for filename in os.listdir(input_dir):
        if filename.endswith('國內個人授信-房貸擔保品.xlsx'):
            file_path = os.path.join(input_dir, filename)
            xls = pd.ExcelFile(file_path)

            pd_rate, pd_rate_orderly_2030, pd_rate_disorderly_2030, pd_rate_no_policy_2030, pd_rate_orderly_2050, pd_rate_disorderly_2050, pd_rate_no_policy_2050, pd_rate_no_policy_2090 = process_PD_domestic_personal_mortgage(file_path)
            lgd_rate, lgd_rate_orderly_2030, lgd_rate_disorderly_2030, lgd_rate_no_policy_2030, lgd_rate_orderly_2050, lgd_rate_disorderly_2050, lgd_rate_no_policy_2050, lgd_rate_no_policy_2090 = process_LGD_domestic_personal_mortgage(file_path)
            ead = process_EAD_domestic_personal_mortgage(file_path)

            summary_2030 = []
            summary_2050 = []
            
            for sheet_name in xls.sheet_names:
                df = pd.read_excel(xls, sheet_name=sheet_name)
                summary_2030.append(df)
                summary_2050.append(df)
            
            summary_df_2030 = pd.concat(summary_2030, ignore_index=True)
            summary_df_2050 = pd.concat(summary_2050, ignore_index=True)

            output_file_path = os.path.join(output_dir, filename)
            with pd.ExcelWriter(output_file_path, engine='openpyxl') as writer:
                save_to_excel(summary_df_2030, '國內授信彙總表(2030年)(個人授信-房貸擔保品)', writer, pd_rate, lgd_rate, pd_rate_orderly_2030, lgd_rate_orderly_2030, pd_rate_disorderly_2030, lgd_rate_disorderly_2030, pd_rate_no_policy_2030, lgd_rate_no_policy_2030, ead)
                save_to_excel(summary_df_2050, '國內授信彙總表(2050年)(個人授信-房貸擔保品)', writer, pd_rate, lgd_rate, pd_rate_orderly_2050, lgd_rate_orderly_2050, pd_rate_disorderly_2050, lgd_rate_disorderly_2050, pd_rate_no_policy_2050, lgd_rate_no_policy_2050, ead)

                print(f"結果已儲存至 {output_file_path}")

def process_domestic_personal_other_credit():
    input_dir = 'input-data_files/授信部位/國內授信/'
    output_dir = 'results-data_files/授信部位/國內授信/'

    if not os.path.exists(output_dir):
        os.makedirs(output_dir)

    for filename in os.listdir(input_dir):
        if filename.endswith('國內個人授信-其他擔保品.xlsx'):
            file_path = os.path.join(input_dir, filename)
            xls = pd.ExcelFile(file_path)

            pd_rate, pd_rate_orderly_2030, pd_rate_disorderly_2030, pd_rate_no_policy_2030, pd_rate_orderly_2050, pd_rate_disorderly_2050, pd_rate_no_policy_2050, pd_rate_no_policy_2090 = process_PD_domestic_personal_other(file_path)
            lgd_rate, lgd_rate_orderly_2030, lgd_rate_disorderly_2030, lgd_rate_no_policy_2030, lgd_rate_orderly_2050, lgd_rate_disorderly_2050, lgd_rate_no_policy_2050, lgd_rate_no_policy_2090 = process_LGD_domestic_personal_other(file_path)
            ead = process_EAD_domestic_personal_other(file_path)

            summary_2030 = []
            summary_2050 = []
            
            for sheet_name in xls.sheet_names:
                df = pd.read_excel(xls, sheet_name=sheet_name)
                summary_2030.append(df)
                summary_2050.append(df)
            
            summary_df_2030 = pd.concat(summary_2030, ignore_index=True)
            summary_df_2050 = pd.concat(summary_2050, ignore_index=True)

            output_file_path = os.path.join(output_dir, filename)
            with pd.ExcelWriter(output_file_path, engine='openpyxl') as writer:
                save_to_excel(summary_df_2030, '國內授信彙總表(2030年)(個人授信-其他擔保品)', writer, pd_rate, lgd_rate, pd_rate_orderly_2030, lgd_rate_orderly_2030, pd_rate_disorderly_2030, lgd_rate_disorderly_2030, pd_rate_no_policy_2030, lgd_rate_no_policy_2030, ead)
                save_to_excel(summary_df_2050, '國內授信彙總表(2050年)(個人授信-其他擔保品)', writer, pd_rate, lgd_rate, pd_rate_orderly_2050, lgd_rate_orderly_2050, pd_rate_disorderly_2050, lgd_rate_disorderly_2050, pd_rate_no_policy_2050, lgd_rate_no_policy_2050, ead)

                print(f"結果已儲存至 {output_file_path}")

def process_overseas_credit():
    input_dir = 'input-data_files/授信部位/國外授信/'
    output_dir = 'results-data_files/授信部位/國外授信/'

    if not os.path.exists(output_dir):
        os.makedirs(output_dir)

    for filename in os.listdir(input_dir):
        if filename.endswith('國外授信.xlsx'):
            file_path = os.path.join(input_dir, filename)
            xls = pd.ExcelFile(file_path)

            pd_rate, pd_rate_orderly_2030, pd_rate_orderly_2050, pd_rate_disorderly_2030, pd_rate_disorderly_2050 = process_PD_overseas_credit(file_path)
            lgd_rate, lgd_rate_orderly_2030, lgd_rate_disorderly_2030, lgd_rate_orderly_2050, lgd_rate_disorderly_2050 = process_LGD_overseas_credit(file_path)
            ead = process_EAD_overseas_credit(file_path)

            summary_2030 = []
            summary_2050 = []
            
            for sheet_name in xls.sheet_names:
                df = pd.read_excel(xls, sheet_name=sheet_name)
                summary_2030.append(df)
                summary_2050.append(df)
            
            summary_df_2030 = pd.concat(summary_2030, ignore_index=True)
            summary_df_2050 = pd.concat(summary_2050, ignore_index=True)

            output_file_path = os.path.join(output_dir, filename)
            with pd.ExcelWriter(output_file_path, engine='openpyxl') as writer:
                save_to_excel_overseas(summary_df_2030, '國外授信彙總表(2030年)', writer, pd_rate, lgd_rate, pd_rate_orderly_2030, lgd_rate_orderly_2030, pd_rate_disorderly_2030, lgd_rate_disorderly_2030, ead)
                save_to_excel_overseas(summary_df_2050, '國外授信彙總表(2050年)', writer, pd_rate, lgd_rate, pd_rate_orderly_2050, lgd_rate_orderly_2050, pd_rate_disorderly_2050, lgd_rate_disorderly_2050, ead)

                print(f"結果已儲存至 {output_file_path}")

def process_domestic_investment():
    input_dir = 'input-data_files/投資部位/國內投資/'
    output_dir = 'results-data_files/投資部位/國內投資/'

    if not os.path.exists(output_dir):
        os.makedirs(output_dir)

    for filename in os.listdir(input_dir):
        if filename.endswith('.xlsx'):
            file_path = os.path.join(input_dir, filename)
            xls = pd.ExcelFile(file_path)

            pd_rate, pd_rate_orderly_2030, pd_rate_disorderly_2030, pd_rate_no_policy_2030, pd_rate_orderly_2050, pd_rate_disorderly_2050, pd_rate_no_policy_2050, pd_rate_no_policy_2090 = process_PD_domestic_investment(file_path)
            if filename.endswith('股權投資部位.xlsx'):
                lgd_rate = [{"客戶名": row['客戶名'], "情境": scenario, "LGD(%)": 1} for row in pd.read_excel(file_path).to_dict('records') for scenario in ["基準情境", "2050淨零轉型 2030", "無序轉型 2030", "無政策情境 2030", "2050淨零轉型 2050", "無序轉型 2050", "無政策情境 2050", "無政策情境 2090"]]
                lgd_rate_orderly_2030 = lgd_rate_disorderly_2030 = lgd_rate_no_policy_2030 = lgd_rate_orderly_2050 = lgd_rate_disorderly_2050 = lgd_rate_no_policy_2050 = lgd_rate_no_policy_2090 = lgd_rate
            else:
                lgd_rate, lgd_rate_orderly_2030, lgd_rate_disorderly_2030, lgd_rate_no_policy_2030, lgd_rate_orderly_2050, lgd_rate_disorderly_2050, lgd_rate_no_policy_2050, lgd_rate_no_policy_2090 = process_LGD_domestic_investment(file_path)

            ead = process_EAD_domestic_investment(file_path)

            summary_2030 = []
            summary_2050 = []

            for sheet_name in xls.sheet_names:
                df = pd.read_excel(xls, sheet_name=sheet_name)
                summary_2030.append(df)
                summary_2050.append(df)

                summary_df_2030 = pd.concat(summary_2030, ignore_index=True)
                summary_df_2050 = pd.concat(summary_2050, ignore_index=True)

                output_file_path = os.path.join(output_dir, filename)
                with pd.ExcelWriter(output_file_path, engine='openpyxl') as writer:
                    save_to_excel(summary_df_2030, '國內投資彙總表(2030年)', writer, pd_rate, lgd_rate, pd_rate_orderly_2030, lgd_rate_orderly_2030, pd_rate_disorderly_2030, lgd_rate_disorderly_2030, pd_rate_no_policy_2030, lgd_rate_no_policy_2030, ead)
                    save_to_excel(summary_df_2050, '國內投資彙總表(2050年)', writer, pd_rate, lgd_rate, pd_rate_orderly_2050, lgd_rate_orderly_2050, pd_rate_disorderly_2050, lgd_rate_disorderly_2050, pd_rate_no_policy_2050, lgd_rate_no_policy_2050, ead)

                print(f"結果已儲存至 {output_file_path}")

def process_overseas_investment():
    input_dir = 'input-data_files/投資部位/國外投資/'
    output_dir = 'results-data_files/投資部位/國外投資/'

    if not os.path.exists(output_dir):
        os.makedirs(output_dir)

    for filename in os.listdir(input_dir):
        if filename.endswith('國外投資.xlsx'):
            file_path = os.path.join(input_dir, filename)
            xls = pd.ExcelFile(file_path)

            pd_rate, pd_rate_orderly_2030, pd_rate_orderly_2050, pd_rate_disorderly_2030, pd_rate_disorderly_2050 = process_PD_overseas_investment(file_path)
            lgd_rate, lgd_rate_orderly_2030, lgd_rate_disorderly_2030, lgd_rate_orderly_2050, lgd_rate_disorderly_2050 = process_LGD_overseas_investment(file_path)
            ead = process_EAD_overseas_investment(file_path)

            summary_2030 = []
            summary_2050 = []
            
            for sheet_name in xls.sheet_names:
                df = pd.read_excel(xls, sheet_name=sheet_name)
                summary_2030.append(df)
                summary_2050.append(df)
            
            summary_df_2030 = pd.concat(summary_2030, ignore_index=True)
            summary_df_2050 = pd.concat(summary_2050, ignore_index=True)

            output_file_path = os.path.join(output_dir, filename)
            with pd.ExcelWriter(output_file_path, engine='openpyxl') as writer:
                save_to_excel_overseas(summary_df_2030, '國外投資彙總表(2030年)', writer, pd_rate, lgd_rate, pd_rate_orderly_2030, lgd_rate_orderly_2030, pd_rate_disorderly_2030, lgd_rate_disorderly_2030, ead)
                save_to_excel_overseas(summary_df_2050, '國外投資彙總表(2050年)', writer, pd_rate, lgd_rate, pd_rate_orderly_2050, lgd_rate_orderly_2050, pd_rate_disorderly_2050, lgd_rate_disorderly_2050, ead)

                print(f"結果已儲存至 {output_file_path}")


def main():
    process_domestic_corporate_credit()
    process_domestic_personal_mortgage_credit()
    process_domestic_personal_other_credit()
    process_overseas_credit()
    process_domestic_investment()
    process_overseas_investment()
if __name__ == "__main__":
    main()
